import openpyxl


class HomePageData:
    test_HomePage_data = {"firstname": "A", "email": "szybkibil@home.pl", "gender": "Male"}
    my_word_list = [
        'danish123', 'cheese%$cake', 'Sugar',
        'Lollipop%', 'Wafer', 'Gu12mmies',
        'sesL7ame', 'Jelly12%!', 'beans&',
        'Pie', 'bardfg986', 'Ic75e', 'oat67']

    @staticmethod
    def getTestData(test_case_name):
        Dict = {}
        book = openpyxl.load_workbook(
            "C:\\Users\\DELL\\PycharmProjects\\PythonSelFramework\\PythonSelFramework\\TestData\\PythonDemo.xlsx")
        sheet = book.active
        for i in range(1, sheet.max_row + 1):  # to get rows
            if sheet.cell(row=i, column=1).value == test_case_name:

                for j in range(2, sheet.max_column + 1):  # to get columns
                    # Dict["email"]="anna@onet.pl"
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return [Dict]
